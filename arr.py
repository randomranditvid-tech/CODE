import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# =========================================================
# USAGE
# =========================================================
#
# python rearrange_risk.py <input_folder> <output_folder>
#
# Example:
#
# python rearrange_risk.py "C:\input" "C:\output"
#
# =========================================================

if len(sys.argv) != 3:

    print("\nUsage:")
    print('python rearrange_risk.py <input_folder> <output_folder>')

    sys.exit(1)

INPUT_FOLDER = sys.argv[1]
OUTPUT_FOLDER = sys.argv[2]

TARGET_SHEET = "Vulnerability Details"

# =========================================================
# VALIDATE PATHS
# =========================================================

if not os.path.exists(INPUT_FOLDER):

    print(f"[ERROR] Input folder not found: {INPUT_FOLDER}")
    sys.exit(1)

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# =========================================================
# COLORS
# =========================================================

HIGH_FILL = PatternFill(
    start_color="FF0000",
    end_color="FF0000",
    fill_type="solid"
)

MEDIUM_FILL = PatternFill(
    start_color="FFA500",
    end_color="FFA500",
    fill_type="solid"
)

LOW_FILL = PatternFill(
    start_color="FFFF00",
    end_color="FFFF00",
    fill_type="solid"
)

BLACK_FONT = Font(color="000000")

# =========================================================
# RISK PRIORITY
# =========================================================

risk_priority = {
    "high": 1,
    "medium": 2,
    "low": 3
}

# =========================================================
# PROCESS FILES
# =========================================================

excel_files = [
    f for f in os.listdir(INPUT_FOLDER)
    if f.lower().endswith(".xlsx")
]

print(f"[+] Found {len(excel_files)} Excel files")

for file_name in excel_files:

    input_file = os.path.join(INPUT_FOLDER, file_name)
    output_file = os.path.join(OUTPUT_FOLDER, file_name)

    print(f"\n=================================================")
    print(f"[+] Processing: {file_name}")

    try:

        wb = load_workbook(input_file)

        if TARGET_SHEET not in wb.sheetnames:

            print(f"[-] Sheet not found: {TARGET_SHEET}")
            continue

        ws = wb[TARGET_SHEET]

        # =================================================
        # FIND HEADER ROW
        # =================================================

        header_row = None
        headers = {}

        for row in ws.iter_rows(min_row=1, max_row=20):

            for cell in row:

                if cell.value:

                    value = str(cell.value).strip()

                    if value.lower() == "risk level":

                        header_row = cell.row
                        break

            if header_row:
                break

        if not header_row:

            print("[-] Could not find 'Risk Level'")
            continue

        # =================================================
        # MAP HEADERS
        # =================================================

        for cell in ws[header_row]:

            if cell.value:

                headers[
                    str(cell.value).strip()
                ] = cell.column

        if "Risk Level" not in headers:

            print("[-] Risk Level column missing")
            continue

        risk_col = headers["Risk Level"]

        # =================================================
        # READ DATA ROWS
        # =================================================

        data_rows = []

        for row_num in range(header_row + 1, ws.max_row + 1):

            row_data = []

            for col_num in range(1, ws.max_column + 1):

                row_data.append(
                    ws.cell(row=row_num, column=col_num).value
                )

            risk_value = ws.cell(
                row=row_num,
                column=risk_col
            ).value

            if risk_value:

                risk_value = str(risk_value).strip().lower()

            else:

                risk_value = "low"

            priority = risk_priority.get(risk_value, 999)

            data_rows.append((priority, risk_value, row_data))

        # =================================================
        # SORT ROWS
        # =================================================

        data_rows.sort(key=lambda x: x[0])

        # =================================================
        # CLEAR OLD DATA
        # =================================================

        for row_num in range(header_row + 1, ws.max_row + 1):

            for col_num in range(1, ws.max_column + 1):

                ws.cell(
                    row=row_num,
                    column=col_num
                ).value = None

                ws.cell(
                    row=row_num,
                    column=col_num
                ).fill = PatternFill(fill_type=None)

                ws.cell(
                    row=row_num,
                    column=col_num
                ).font = Font(color="000000")

        # =================================================
        # WRITE SORTED DATA
        # =================================================

        current_row = header_row + 1

        for _, risk_value, row_data in data_rows:

            for col_num, value in enumerate(row_data, start=1):

                cell = ws.cell(
                    row=current_row,
                    column=col_num
                )

                cell.value = value

                # =========================================
                # APPLY COLORS
                # =========================================

                if risk_value == "high":

                    cell.fill = HIGH_FILL
                    cell.font = BLACK_FONT

                elif risk_value == "medium":

                    cell.fill = MEDIUM_FILL
                    cell.font = BLACK_FONT

                elif risk_value == "low":

                    cell.fill = LOW_FILL
                    cell.font = BLACK_FONT

            current_row += 1

        # =================================================
        # SAVE FILE
        # =================================================

        wb.save(output_file)

        print(f"[+] Saved: {output_file}")

    except Exception as e:

        print(f"[ERROR] {file_name}")
        print(str(e))

print("\n=================================================")
print("[+] All files processed successfully")